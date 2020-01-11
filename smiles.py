import ifg
import re
from openpyxl import Workbook

# # Openpyxl objects
# fgWorkbook = Workbook()
# fgSheet = fgWorkbook.active
#
# # Set intial column names
# fgSheet.cell(row=1, column=1).value = "Refcode"
# fgSheet.cell(row=1, column=2).value = "SMILES"
#
#
# # List which holds functional group names
# functionalGroups = []
#
# # Grab all names from FGlist.txt
# for line in open('FGlist.txt', 'r'):
#     lineInfo = re.compile(r'\S+').findall(line)
#     groupName = lineInfo[1]
#     if groupName not in functionalGroups:
#         functionalGroups.append(groupName)
# functionalGroups.append("Alcohol")
#
# # Sort them alphabetically
# functionalGroups.sort()
# holder = functionalGroups.copy()
#
# # Add cyclic/aromatic distinctions to list
# indexCounter = insertions = 0
# for name in holder:
#     indexCounter += 1
#     functionalGroups.insert(indexCounter+insertions, "Cyclic" + name)
#     if len(re.compile(r'Amine').findall(name)) != 0:
#         functionalGroups.insert(indexCounter+insertions, "Aromatic" + name)
#         insertions += 1
#     insertions += 1
#
# functionalGroups.pop(3) # Get rid of cyclic alcohol
#
#
# # Insert names into the top row
# for column in range(3, len(functionalGroups)+3):
#     fgSheet.cell(row=1,column=column).value = functionalGroups[column-3]
#
# fgSheet.cell(row=1, column=len(functionalGroups)+3).value = "aromaticRingCount"
# fgSheet.cell(row=1, column=len(functionalGroups)+4).value = "nonAromaticRingCount"
# fgSheet.cell(row=1, column=len(functionalGroups)+5).value = "RingCount"
# fgSheet.cell(row=1, column=len(functionalGroups)+6).value = "AminoAcid"
#
# rowCounter = 1
# maxColumn = fgSheet.max_column
# print(functionalGroups)
#
# codes = []
for line in open('smiles.txt', 'r'):
# # # #     rowCounter += 1
# # # #     for column in range(2, maxColumn+1):
# # # #         fgSheet.cell(row=rowCounter, column=column).value = 0
    lineInfo = re.compile(r'\S+').findall(line)
    smiles = lineInfo[1]
    RefCode = lineInfo[2]
# # # #     fgSheet.cell(row=rowCounter, column=1).value = RefCode
# # # #     fgSheet.cell(row=rowCounter, column=2).value = smilesWOH
    print("EVALUATING ", lineInfo[2], " ", smiles)
    functionalGroupData = ifg.ifg(smiles)
    print(functionalGroupData)
# # #     for group in functionalGroupData[1].items():
# # #         for column in range(2,maxColumn+1):
# # #             if fgSheet.cell(row=1, column=column).value == group[0]:
# # #                 fgSheet.cell(row=rowCounter, column=column).value = int(group[1])
# # #                 if group[0] == "AminoAcid" and int(group[1]) == 1:
# # #                     fgSheet.cell(row=rowCounter, column=column).value = "Yes"
# # #                 elif group[0] == "AminoAcid" and int(group[1]) == 0:
# # #                     fgSheet.cell(row=rowCounter, column=column).value = "No"
# #                 break
    del(functionalGroupData)
# fgWorkbook.save("functionalGroupData.xlsx")

# data = ifg.ifg("Cc1c(N=[N+]=[N-])nc(N=[N+]=[N-])nc1N=[N+]=[N-]")
# print("Cc1c(N=[N+]=[N-])nc(N=[N+]=[N-])nc1N=[N+]=[N-]")
# print(data)
